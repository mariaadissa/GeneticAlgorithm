# -*- coding: utf-8 -*-
"""
Created on Sat May  8 00:44:57 2021

@author: maria
"""
import openpyxl
import numpy as np
import operator
from random import randint, randrange
import random
import time


class Bin(object):
    """ This class implements the bin object """
    
    def __init__(self):
        self.items = []  #the list that is going to contain the items that fit in the bin
        self.total_weight = 0  #integer that stores the total weight of the items in the bin

    def add_item(self, item): #this function adds an item in the bin
        self.items.append(item)
        self.total_weight += item


class Chromosome:
    """This class implements a candidate solution"""
    
    def __init__(self):
        self.genome = []  #the list containing the items in a randomly generated order
        self.fitness = 0  #int stores the number of bins used in this solution (should be minimized)
    
    def add_genome(self, list): #this function adds an item in the bin
        self.genome=list
        
        

def read(f):
    """This function reads a given file and returns a tuple 
    that has the list containing the items and an integer=the bin capacity"""
    
    book = openpyxl.load_workbook(f)  
    sheet = book.active      
    items_list =[]
    n = sheet.cell(row = 2, column = 1).value
    c = sheet.cell(row = 2, column = 2).value
    for i in range(2,n+2):  
        cell_obj = sheet.cell(row = i, column = 3)
        items_list.append(cell_obj.value) 
    return items_list,c  


def ffa(items_list, bin_capacity):
    bins =[]
        
    for item in items_list:
        # foeach item we search if there's an open bin where it can fit
        for bin in bins:
            if bin.total_weight + item <= bin_capacity: #if it fits
                bin.add_item(item)  #we add the item in the bin
                break
        else:
            # there is no open bin where the item can fit
            #so we open a new bin and add the item in it
            bin = Bin()
            bin.add_item(item)
            bins.append(bin)
 
    return bins

""" Generating the population
"""
def generate_population(k,items_list):
    """This function generates a population of k candidates from the initial items list"""
    
    population=[]
    randomised_np_list=[]
    randomised_list=[]
    for i in range(k):
        randomised_np_list = np.random.permutation(items_list)
        randomised_list = randomised_np_list.tolist()
        chromosome = Chromosome()
        chromosome.add_genome(randomised_list)
        population.append(chromosome)
        
    return population

""" 1st STEP: SELECTION
"""

def calculate_fitness(chromosome: Chromosome,c):
    """This function calculates the fitness of a given chromosome and given capacity of the bins used"""
    
    bins = ffa(chromosome.genome,c)  #apply the ffa algorithm on the genome
    chromosome.fitness=len(bins)     #chromosome fitness is the number of bins used by ffa


def sort_population(population: list):
    """Returns the sorted population of chromosomes based on fitness value"""
    return sorted(population, key=operator.attrgetter('fitness'))

  
def selection(population: list,bin_capacity):
    """returns two best candidates(parents)"""
    for chromosome in population:
        calculate_fitness(chromosome, bin_capacity)
    sorted_population=sort_population(population)
    return sorted_population[0],sorted_population[1] #return two candidatses with smallest fitness 
    
""" 2nd STEP: CROSSOVER (single point crossover)
"""
def single_point_crossover(parent1: Chromosome,parent2: Chromosome):
    """Implements the single point crossover method on the two chosen chromosomes"""
    
    genome1 = parent1.genome
    genome2 = parent2.genome
    
    if len(genome1) != len(genome2):
        raise ValueError("Genomes 1 and 2 must be of same length!")

    length = len(genome1)
    p = randint(1, length - 1)  #random point of the crossover
    child=Chromosome()
    child.add_genome(genome1[0:p] + genome2[p:]) #1st part of genome1 is concatenated with 2nd part of genome2
    return child

""" 3rd STEP: MUTATION
"""
def mutation(probability,child: Chromosome):
    """This function returns a new child if the probability is higher than the chosen probability"""
    
    genome = child.genome
    p=random.uniform(0, 1) #generating a random number between 0 and 1
    index1 = randrange(len(genome))
    index2 = randrange(len(genome))
    
    if p>probability: #if the generated number is greater than the probability
        return -1
    else:
        genome[index1], genome[index2] = genome[index2], genome[index1] #permutate the index1 and index2 orders
        chromosome = Chromosome()
        chromosome.add_genome(genome)
        return chromosome 

""" 4th STEP: UPDATING POPULATION
"""
def population_update(population: list,new_child,bin_capacity):
    """Updates the population with the new child from mutation process and deletes worst candidate from population"""
    if isinstance(new_child,int):  #if there is no mutation
         print('no mutation')
    else:
         population.append(new_child)
         calculate_fitness(new_child,bin_capacity)
    new_population=sort_population(population)
    new_population.pop()   #removing the chromosome with highest fitness from population
    return new_population
    
def best_solution(population):
    return population[0]
        
def genetic_algo(k,items_list,bin_capacity,max_iteration,optimal_solution):
    population = generate_population(k,items_list)
    iteration = 0
    while(iteration<max_iteration):
        iteration= iteration + 1
        #Selection
        parent1,parent2 = selection(population,bin_capacity)
        #Crossover
        child = single_point_crossover(parent1,parent2)
        #Mutation
        new_child = mutation(MUTATION_POBABILITY,child)
        #Population update
        new_population= population_update(population,new_child,bin_capacity)
        best_solution_found = best_solution(new_population)
        if best_solution_found.fitness== optimal_solution:
            return best_solution_found
    return best_solution(new_population)    
        


if __name__ == '__main__': 
     
     result=read("HARD0.xlsx")
     
     items_list,c = result[0],result[1] 
     
     MUTATION_POBABILITY = 0.1  
     
     N1C1W1_A_solution = 25
     N1C1W1_B_solution = 31
     HARD0_solution = 56
     
     start_time_genetic = time.time()
     solution_genetic=genetic_algo(50, items_list, c, 100, N1C1W1_A_solution)
     
     print(solution_genetic.fitness)
     print("%s seconds " % (time.time() - start_time_genetic))
     
     ##FFA Algorithm execution
     start_time_ffa = time.time()
     solution_ffa = len(ffa(items_list,c))
     
     print(solution_ffa)
     print("%s seconds " % (time.time() - start_time_ffa))
     